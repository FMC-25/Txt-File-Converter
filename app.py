import streamlit as st
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

st.title("MT54x to Excel Converter")

# ── 1. Parse MT54x messages ───────────────────────────────────────────────────
def parse_mt54x(text):
    messages = re.findall(r'\{4:(.*?)-\}', text, re.DOTALL)
    records = []
    for msg in messages:
        def get(pattern):
            m = re.search(pattern, msg)
            return m.group(1).strip() if m else ''

        seme      = get(r':20C::SEME//([^\n]+)')
        sett_date = get(r':98A::SETT//(\d{8})')
        trad_date = get(r':98A::TRAD//(\d{8})')
        isin      = get(r':35B:ISIN ([^\n]+)')
        face_amt  = get(r':36B::SETT//FAMT/([\d,]+)')
        sett_amt  = get(r':19A::SETT//LKR([\d,]+)')

        counterparty = ''

        fiac_safe = re.search(
            r':16R:FIAC.*?:97A::SAFE//(?!OWN)([^\n]+).*?:16S:FIAC', msg, re.DOTALL
        )
        if fiac_safe:
            raw = fiac_safe.group(1).strip()
            counterparty = re.sub(
                r'^(RETE|RETNO|RETGAM|RRT|RET|CSL)', '', raw, flags=re.IGNORECASE
            ).strip()
        elif re.search(r':95P::REAG//NSBFLKLX.*?:97A::SAFE//(?!OWN)', msg, re.DOTALL):
            reag_block = re.search(
                r':95P::REAG//NSBFLKLX.*?:97A::SAFE//([^\n]+)', msg, re.DOTALL
            )
            if reag_block:
                raw = reag_block.group(1).strip()
                counterparty = re.sub(
                    r'^(CSL|RET)', '', raw, flags=re.IGNORECASE
                ).strip()
        else:
            deag = re.search(r':95P::DEAG//([^\n]+)', msg)
            reag = re.search(r':95P::REAG//(?!NSBFLKLX)([^\n]+)', msg)
            if deag:
                counterparty = deag.group(1).strip()
            elif reag:
                counterparty = reag.group(1).strip()

        def fmt_date(d):
            return f"{d[:4]}-{d[4:6]}-{d[6:]}" if len(d) == 8 else d

        def to_int(s):
            try:    return int(s.replace(',', '').replace('.', ''))
            except ValueError: return s

        records.append({
            'Message Ref (SEME)':     seme,
            'Trade Date':             fmt_date(trad_date),
            'Settlement Date':        fmt_date(sett_date),
            'ISIN':                   isin,
            'Face Amount (LKR)':      to_int(face_amt),
            'Counterparty / Account': counterparty,
            'Settlement Amt (LKR)':   to_int(sett_amt),
            'Verification':           '',
            'Authorization':          '',
        })
    return records

def get_settlement_date(records):
    return records[0].get('Settlement Date', '') if records else ''

# ── 2. Styling helpers ────────────────────────────────────────────────────────
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
TOTAL_FILL  = PatternFill('solid', start_color='BDD7EE')
FILL_LIGHT  = PatternFill('solid', start_color='DCE6F1')
FILL_WHITE  = PatternFill('solid', start_color='FFFFFF')
DATA_FONT   = Font(name='Arial', size=10)
BOLD_FONT   = Font(name='Arial', bold=True)
THIN        = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
AMT_FMT = '#,##0'

def style_cell(cell, font=None, fill=None, number_format=None, align=None):
    if isinstance(cell, MergedCell):
        return
    cell.border    = THIN
    cell.alignment = align or Alignment(horizontal='center', vertical='center')
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if number_format: cell.number_format = number_format

# ── 3. Write individual sheet ─────────────────────────────────────────────────
def write_sheet(ws, records, sheet_label):
    if not records:
        ws['A1'] = 'No records found'
        return

    headers   = list(records[0].keys())
    num_cols  = len(headers)
    last_col  = get_column_letter(num_cols)
    sett_date = get_settlement_date(records)

    ws.merge_cells(f'A1:{last_col}1')
    date_cell           = ws['A1']
    date_cell.value     = f"Settlement Date: {sett_date}"
    date_cell.font      = Font(name='Arial', bold=True, size=12, color='1F4E79')
    date_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    for col_idx, h in enumerate(headers, 1):
        style_cell(ws.cell(2, col_idx, h), font=HEADER_FONT, fill=HEADER_FILL)

    for row_idx, rec in enumerate(records, 3):
        row_fill = FILL_LIGHT if row_idx % 2 == 0 else FILL_WHITE
        for col_idx, value in enumerate(rec.values(), 1):
            cell = ws.cell(row_idx, col_idx, value)
            fmt  = AMT_FMT if col_idx in (5, 7) else None
            style_cell(cell, font=DATA_FONT, fill=row_fill, number_format=fmt)

    sig_row = ws.max_row + 3
    sig_font = Font(name='Arial', size=11)

    ws.merge_cells(f'A{sig_row}:D{sig_row}')
    verified_cell           = ws[f'A{sig_row}']
    verified_cell.value     = 'Verified by:  ' + '.' * 40
    verified_cell.font      = sig_font
    verified_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[sig_row].height = 24

    auth_row = sig_row + 2
    ws.merge_cells(f'A{auth_row}:D{auth_row}')
    auth_cell           = ws[f'A{auth_row}']
    auth_cell.value     = 'Authorized by:' + '.' * 40
    auth_cell.font      = sig_font
    auth_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[auth_row].height = 24

    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row, col_idx).value))
             for row in range(2, ws.max_row + 1)
             if ws.cell(row, col_idx).value
             and not isinstance(ws.cell(row, col_idx), MergedCell)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    ws.freeze_panes = 'A3'

# ── 4. Main Application Logic ─────────────────────────────────────────────────
uploaded_files = st.file_uploader("Upload MT541/542/543 .txt files", type=['txt'], accept_multiple_files=True)

if uploaded_files:
    if st.button("Process Files"):
        wb = Workbook()
        wb.remove(wb.active)

        # Sort files by name
        uploaded_files.sort(key=lambda x: x.name)

        # Process individual files
        for uploaded_file in uploaded_files:
            content = uploaded_file.read().decode('utf-8', errors='replace')
            records = parse_mt54x(content)
            raw_name = uploaded_file.name.replace('.txt', '').replace('_', ' ')
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', raw_name)[:31]
            ws = wb.create_sheet(title=sheet_name)
            write_sheet(ws, records, sheet_name)

        # ── Summary sheet ──
        ws_sum = wb.create_sheet(title='SUMMARY', index=0)
        sum_headers = ['Sheet / File', 'Settlement Date', 'Transactions']

        ws_sum.append([''])
        for col_idx, h in enumerate(sum_headers, 1):
            style_cell(ws_sum.cell(2, col_idx, h), font=HEADER_FONT, fill=HEADER_FILL)

        data_sheets = [s for s in wb.worksheets if s.title != 'SUMMARY']
        for i, sheet in enumerate(data_sheets, 3):
            # Read content again for the summary
            file_obj = uploaded_files[i - 3]
            file_obj.seek(0)
            content   = file_obj.read().decode('utf-8', errors='replace')
            records   = parse_mt54x(content)
            sett_date = get_settlement_date(records)

            ws_sum.cell(i, 1, sheet.title)
            ws_sum.cell(i, 2, sett_date)
            ws_sum.cell(i, 3, len(records))
            fill = FILL_LIGHT if i % 2 == 0 else FILL_WHITE
            for col_idx in range(1, 4):
                style_cell(ws_sum.cell(i, col_idx), font=DATA_FONT, fill=fill)

        # Grand total row
        grand_row = 3 + len(data_sheets)
        ws_sum.cell(grand_row, 1, 'GRAND TOTAL')
        ws_sum.cell(grand_row, 3, f'=SUM(C3:C{grand_row-1})')
        for col_idx in range(1, 4):
            style_cell(ws_sum.cell(grand_row, col_idx), font=BOLD_FONT, fill=TOTAL_FILL)

        # Summary column widths
        for col_idx in range(1, 4):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(ws_sum.cell(row, col_idx).value))
                 for row in range(2, ws_sum.max_row + 1)
                 if ws_sum.cell(row, col_idx).value
                 and not isinstance(ws_sum.cell(row, col_idx), MergedCell)),
                default=10
            )
            ws_sum.column_dimensions[col_letter].width = min(max_len + 4, 40)
        ws_sum.freeze_panes = 'A3'

        # ── Save to memory and download ──
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)

        st.success(f"Successfully processed {len(uploaded_files)} files.")
        st.download_button(
            label="Download Excel File",
            data=excel_data,
            file_name="MT54x_All_Transactions.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )